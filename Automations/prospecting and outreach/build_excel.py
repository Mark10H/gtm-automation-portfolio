"""
build_excel.py — Moxo BDR Outreach Excel Builder
=================================================
Generates a 7-sheet Excel file from a structured JSON input file.

USAGE:
    python scripts/build_excel.py <input_json_path> <output_dir>

    input_json_path : path to the JSON file Claude produces from session data
    output_dir      : folder where the .xlsx will be saved
                      (defaults to mnt/outputs/ if omitted)

OUTPUT FILENAME:
    Moxo_Outreach_<RepFirstName>_<YYYY-MM-DD>.xlsx

INPUT JSON SCHEMA (Claude must produce this before calling the script):
{
  "rep_name":  "Spencer Johnson",          // full name
  "date":      "2026-04-08",               // ISO date string
  "companies": [                           // Sheet 1
    {
      "company":           "Acme Corp",
      "website":           "acmecorp.com",
      "industry":          "Financial Services",
      "employees":         250,
      "rep":               "Spencer Johnson",
      "enrichment_score":  87,             // composite 1-100 (ZI 55% + web 45%)
      "zi_sub_score":      48,             // ZoomInfo sub-score (max 55)
      "web_sub_score":     39,             // web research sub-score (max 45)
      "confidence_rating": "High",         // High / Medium / Low
      "why_selected":      "One sentence rationale"
    }
  ],
  "account_intelligence": [                // Sheet 2
    {
      "company":          "Acme Corp",
      "trigger_event":    "...",
      "trigger_category":  "Digital Transformation & AI",
      "evidence_url":      "https://...",
      "process":           "One sentence — what process Moxo supports for this account",
      "how_moxo_helps":    "Max 3 sentences tied to trigger and industry",
      "without_moxo":      "Max 2 sentences — risks/costs of doing nothing",
      "business_outcome":  "Max 2 sentences — real stats from KB",
      "target_personas":   "COO | VP of Operations | ..."
    }
  ],
  "email_sequences": [                     // Sheet 3 (generic, 4 per company)
    {
      "company": "Acme Corp",
      "email_number": 1,
      "type":    "New Thread",
      "subject": "Subject line here",
      "body":    "Email body here"
    }
  ],
  "contacts": [                            // Sheet 4
    {
      "company":      "Acme Corp",
      "first_name":   "Jane",
      "last_name":    "Smith",
      "title":        "VP of Operations",
      "tier":         "Tier 1 — Tablestakes",
      "linkedin_url": "https://linkedin.com/in/...",
      "email":        "jane@acmecorp.com",
      "phone":        "555-555-5555"
    }
  ],
  "contact_sequences": [                   // Sheet 5 (4 emails per contact)
    {
      "company":             "Acme Corp",
      "contact_first_name":  "Jane",
      "contact_last_name":   "Smith",
      "title":               "VP of Operations",
      "tier":                "Tier 1 — Tablestakes",
      "email_number":        1,
      "type":                "New Thread",
      "subject":             "Subject line here",
      "body":                "Email body here",
      "research_notes":      "Tailored — LinkedIn post re: AI ops initiative"
    }
  ],
  "territory_health": {                    // Sheet 6 — optional; omit key or set null to show stub row
    "rep_name":                      "Spencer Johnson",
    "total_accounts_owned":          120,
    "accounts_researched_all_time":  34,
    "accounts_researched_session":   5,
    "accounts_excluded_session":     8,
    "accounts_remaining_eligible":   78,
    "top_exclusion_reason":          "Excluded — Outside Headcount Range",
    "pct_territory_enriched":        "28%",
    "dashboard_note":                ""   // optional free-text note from territory_health.py
  },
  "audit_log": [                           // Sheet 7
    {
      "company_name":       "Acme Corp",
      "website":            "acmecorp.com",
      "industry":           "Financial Services",
      "company_owner":      "Spencer Johnson",
      "status":             "Researched",  // Researched | Excluded
      "trigger_category":   "Digital Transformation & AI",
      "why_selected":       "One sentence",
      "business_impact":    "One sentence",
      "exclusion_reason":   "",            // blank if Researched
      "kb_alignment_note":  ""            // blank unless excluded for no fit
    }
  ]
}
"""

import sys
import json
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
COLORS = {
    "header_bg":   "1A2E4A",   # dark navy — header row background
    "header_fg":   "FFFFFF",   # white text for headers
    "email_1":     "D6EAF8",   # light blue  — Email 1
    "email_2":     "D5F5E3",   # light green — Email 2
    "email_3":     "FEF9E7",   # light yellow — Email 3
    "email_4":     "FDEDEC",   # light pink/orange — Email 4
    "tier_1":      "EBF5FB",   # very light blue — Tier 1 contacts
    "tier_2":      "F9F9F9",   # near-white — Tier 2 contacts
    "researched":  "EAFAF1",   # light green — Researched rows in audit log
    "excluded":    "FDFEFE",   # off-white — Excluded rows in audit log
    "alt_row":     "F2F3F4",   # light grey — alternating row tint
}

EMAIL_ROW_COLORS = {
    1: COLORS["email_1"],
    2: COLORS["email_2"],
    3: COLORS["email_3"],
    4: COLORS["email_4"],
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_header_font() -> Font:
    return Font(bold=True, color=COLORS["header_fg"], name="Calibri", size=11)


def make_body_font(bold=False) -> Font:
    return Font(bold=bold, name="Calibri", size=10)


def thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def wrap_align(horizontal="left") -> Alignment:
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)


def style_header_row(ws, row_num: int, num_cols: int):
    fill = make_fill(COLORS["header_bg"])
    font = make_header_font()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = wrap_align("center")
        cell.border = thin_border()


def write_headers(ws, headers: list):
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"


def set_col_widths(ws, widths: list):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_cell(ws, row, col, value, fill_color=None, bold=False, row_height=None):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.font = make_body_font(bold=bold)
    cell.alignment = wrap_align()
    cell.border = thin_border()
    if fill_color:
        cell.fill = make_fill(fill_color)
    return cell


def write_row(ws, row_num, values, fill_color=None, bold_cols=None):
    bold_cols = bold_cols or []
    for col_idx, value in enumerate(values, start=1):
        write_cell(ws, row_num, col_idx, value,
                   fill_color=fill_color,
                   bold=(col_idx in bold_cols))


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_sheet1_companies(ws, companies: list):
    """Sheet 1 — Companies Researched"""
    ws.title = "Companies Researched"
    headers = ["Company", "Website", "Industry", "Employees", "Rep",
               "Enrichment Score", "ZI Sub-score", "Web Sub-score",
               "Confidence Rating", "Why Selected"]
    write_headers(ws, headers)
    set_col_widths(ws, [28, 28, 24, 12, 22, 16, 14, 14, 18, 50])

    for row_idx, co in enumerate(companies, start=2):
        fill = COLORS["alt_row"] if row_idx % 2 == 0 else None
        confidence = co.get("confidence_rating", "")
        if confidence == "High":
            conf_fill = "D5F5E3"
        elif confidence == "Medium":
            conf_fill = "FEF9E7"
        elif confidence == "Low":
            conf_fill = "FDEDEC"
        else:
            conf_fill = fill

        # Color-code enrichment score: >=85 green, 70-84 yellow, <70 red (shouldn't appear)
        enrichment_score = co.get("enrichment_score", "")
        if isinstance(enrichment_score, (int, float)):
            if enrichment_score >= 85:
                score_fill = "D5F5E3"   # green - Hot
            elif enrichment_score >= 70:
                score_fill = "FEF9E7"   # yellow - Warm
            else:
                score_fill = "FDEDEC"   # red - below threshold
        else:
            score_fill = fill

        values = [
            co.get("company", ""),
            co.get("website", ""),
            co.get("industry", ""),
            co.get("employees", ""),
            co.get("rep", ""),
            enrichment_score,
            co.get("zi_sub_score", ""),
            co.get("web_sub_score", ""),
            confidence,
            co.get("why_selected", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            if col_idx == 6:
                cell_fill = score_fill
            elif col_idx == 9:
                cell_fill = conf_fill
            else:
                cell_fill = fill
            write_cell(ws, row_idx, col_idx, value, fill_color=cell_fill,
                       bold=(col_idx == 1))
        ws.row_dimensions[row_idx].height = 40


def build_sheet2_intelligence(ws, intelligence: list):
    """Sheet 2 — Account Intelligence"""
    ws.title = "Account Intelligence"
    headers = ["Company", "Process", "Trigger", "Trigger Category",
               "Evidence URL", "How Moxo Helps", "Without Moxo",
               "Business Outcome", "Target Personas"]
    write_headers(ws, headers)
    set_col_widths(ws, [22, 40, 40, 28, 38, 52, 45, 45, 38])

    CATEGORY_COLORS = {
        "News/Milestone":              "D6EAF8",
        "Digital Transformation & AI": "D5F5E3",
        "Operational Gap":             "FEF9E7",
        "Growth/Retention Gap":        "FDEBD0",
    }

    for row_idx, rec in enumerate(intelligence, start=2):
        cat = rec.get("trigger_category", "")
        row_fill = CATEGORY_COLORS.get(cat, COLORS["alt_row"] if row_idx % 2 == 0 else None)
        values = [
            rec.get("company", ""),
            rec.get("process", ""),
            rec.get("trigger", ""),
            cat,
            rec.get("evidence_url", ""),
            rec.get("how_moxo_helps", ""),
            rec.get("without_moxo", ""),
            rec.get("business_outcome", ""),
            rec.get("target_personas", ""),
        ]
        write_row(ws, row_idx, values, fill_color=row_fill, bold_cols=[1])
        ws.row_dimensions[row_idx].height = 80


def build_sheet3_email_sequences(ws, sequences: list):
    """Sheet 3 — Email Sequences (Generic) — color-coded by email number"""
    ws.title = "Email Sequences (Generic)"
    headers = ["Company", "Email #", "Type", "Subject", "Body"]
    write_headers(ws, headers)
    set_col_widths(ws, [22, 10, 18, 45, 90])

    for row_idx, seq in enumerate(sequences, start=2):
        email_num = seq.get("email_number", 1)
        row_fill = EMAIL_ROW_COLORS.get(email_num, None)
        values = [
            seq.get("company", ""),
            email_num,
            seq.get("type", ""),
            seq.get("subject", ""),
            seq.get("body", ""),
        ]
        write_row(ws, row_idx, values, fill_color=row_fill, bold_cols=[1, 2, 4])
        ws.row_dimensions[row_idx].height = 120

    # Legend
    legend_start = len(sequences) + 3
    ws.cell(row=legend_start, column=1, value="COLOR KEY").font = make_body_font(bold=True)
    for num, label, color in [
        (1, "Email 1 — New Thread",    COLORS["email_1"]),
        (2, "Email 2 — Reply (Email 1)", COLORS["email_2"]),
        (3, "Email 3 — Reply (Email 2)", COLORS["email_3"]),
        (4, "Email 4 — Break-up",       COLORS["email_4"]),
    ]:
        row = legend_start + num
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill = make_fill(color)
        cell.font = make_body_font()
        cell.border = thin_border()


def build_sheet4_contacts(ws, contacts: list):
    """Sheet 4 — Contacts"""
    ws.title = "Contacts"
    headers = ["Company", "First Name", "Last Name", "Title",
               "Tier", "LinkedIn URL", "Email", "Phone"]
    write_headers(ws, headers)
    set_col_widths(ws, [22, 14, 16, 34, 24, 42, 30, 18])

    for row_idx, contact in enumerate(contacts, start=2):
        tier = contact.get("tier", "")
        row_fill = COLORS["tier_1"] if "1" in tier else COLORS["tier_2"]
        values = [
            contact.get("company", ""),
            contact.get("first_name", ""),
            contact.get("last_name", ""),
            contact.get("title", ""),
            tier,
            contact.get("linkedin_url", ""),
            contact.get("email", ""),
            contact.get("phone", ""),
        ]
        write_row(ws, row_idx, values, fill_color=row_fill, bold_cols=[1])
        ws.row_dimensions[row_idx].height = 28


def build_sheet5_contact_sequences(ws, contact_sequences: list):
    """Sheet 5 — Contact Email Sequences"""
    ws.title = "Contact Email Sequences"
    headers = ["Company", "First Name", "Last Name", "Title", "Tier",
               "Email #", "Type", "Subject", "Body", "Research Notes"]
    write_headers(ws, headers)
    set_col_widths(ws, [22, 12, 14, 30, 24, 10, 16, 42, 82, 40])

    for row_idx, seq in enumerate(contact_sequences, start=2):
        email_num = seq.get("email_number", 1)
        row_fill = EMAIL_ROW_COLORS.get(email_num, None)
        values = [
            seq.get("company", ""),
            seq.get("contact_first_name", ""),
            seq.get("contact_last_name", ""),
            seq.get("title", ""),
            seq.get("tier", ""),
            email_num,
            seq.get("type", ""),
            seq.get("subject", ""),
            seq.get("body", ""),
            seq.get("research_notes", ""),
        ]
        write_row(ws, row_idx, values, fill_color=row_fill, bold_cols=[1, 6, 8])
        ws.row_dimensions[row_idx].height = 110


def build_sheet6_territory_health(ws, territory: dict):
    """Sheet 6 — Territory Health Dashboard"""
    ws.title = "Territory Health"

    # Title banner
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "Territory Health Dashboard"
    title_cell.font = Font(bold=True, size=13, color=COLORS["header_fg"], name="Calibri")
    title_cell.fill = make_fill(COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    set_col_widths(ws, [38, 32])

    METRICS = [
        ("Rep Name",                        "rep_name"),
        ("Total Accounts Owned",            "total_accounts_owned"),
        ("Accounts Researched (All Time)",  "accounts_researched_all_time"),
        ("Accounts Researched (This Session)", "accounts_researched_session"),
        ("Accounts Excluded (This Session)", "accounts_excluded_session"),
        ("Accounts Remaining (Eligible)",   "accounts_remaining_eligible"),
        ("Top Exclusion Reason",            "top_exclusion_reason"),
        ("% Territory Enriched",            "pct_territory_enriched"),
        ("Notes",                           "dashboard_note"),
    ]

    if not territory:
        # Stub row when territory data wasn't collected this session
        ws.merge_cells("A3:B3")
        stub = ws["A3"]
        stub.value = "Run /territory for full dashboard. Session summary unavailable — territory command was not run this session."
        stub.font = make_body_font()
        stub.alignment = wrap_align()
        stub.fill = make_fill(COLORS["alt_row"])
        ws.row_dimensions[3].height = 40
        return

    for row_idx, (label, key) in enumerate(METRICS, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = make_body_font(bold=True)
        label_cell.alignment = wrap_align()
        label_cell.fill = make_fill(COLORS["alt_row"] if row_idx % 2 == 0 else "FFFFFF")
        label_cell.border = thin_border()

        value = territory.get(key, "")
        value_cell = ws.cell(row=row_idx, column=2, value=str(value) if value is not None else "")
        value_cell.font = make_body_font()
        value_cell.alignment = wrap_align()
        value_cell.fill = make_fill(COLORS["alt_row"] if row_idx % 2 == 0 else "FFFFFF")
        value_cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 22


def build_sheet7_audit_log(ws, audit_log: list):
    """Sheet 7 — Audit Log (Manager)"""
    ws.title = "Audit Log (Manager)"
    headers = [
        "Company Name", "Website", "Industry", "Company Owner",
        "Status", "Trigger Category", "Why We Chose This Account",
        "Business Impact", "Exclusion Reason", "KB Alignment Note"
    ]
    write_headers(ws, headers)
    set_col_widths(ws, [24, 26, 22, 22, 14, 28, 50, 45, 38, 38])

    for row_idx, entry in enumerate(audit_log, start=2):
        status = entry.get("status", "")
        row_fill = COLORS["researched"] if status == "Researched" else COLORS["excluded"]
        values = [
            entry.get("company_name", ""),
            entry.get("website", ""),
            entry.get("industry", ""),
            entry.get("company_owner", ""),
            status,
            entry.get("trigger_category", ""),
            entry.get("why_selected", ""),
            entry.get("business_impact", ""),
            entry.get("exclusion_reason", ""),
            entry.get("kb_alignment_note", ""),
        ]
        write_row(ws, row_idx, values, fill_color=row_fill, bold_cols=[1, 5])
        ws.row_dimensions[row_idx].height = 55


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("USAGE: python scripts/build_excel.py <input_json> [output_dir]")
        print("       input_json  — path to the JSON data file Claude produces")
        print("       output_dir  — folder to save the .xlsx (default: mnt/outputs/)")
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "mnt/outputs"

    # Load data
    if not os.path.exists(input_path):
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Build filename
    rep_name = data.get("rep_name", "Unknown")
    rep_first = rep_name.split()[0] if rep_name else "Rep"
    date_str = data.get("date", datetime.today().strftime("%Y-%m-%d"))
    filename = f"Moxo_Outreach_{rep_first}_{date_str}.xlsx"

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    # Build workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"Building workbook: {filename}")

    ws1 = wb.create_sheet("Companies Researched")
    build_sheet1_companies(ws1, data.get("companies", []))
    print(f"  ✓ Sheet 1 — Companies Researched ({len(data.get('companies', []))} rows)")

    ws2 = wb.create_sheet("Account Intelligence")
    build_sheet2_intelligence(ws2, data.get("account_intelligence", []))
    print(f"  ✓ Sheet 2 — Account Intelligence ({len(data.get('account_intelligence', []))} rows)")

    ws3 = wb.create_sheet("Email Sequences (Generic)")
    build_sheet3_email_sequences(ws3, data.get("email_sequences", []))
    print(f"  ✓ Sheet 3 — Email Sequences (Generic) ({len(data.get('email_sequences', []))} rows)")

    ws4 = wb.create_sheet("Contacts")
    build_sheet4_contacts(ws4, data.get("contacts", []))
    print(f"  ✓ Sheet 4 — Contacts ({len(data.get('contacts', []))} rows)")

    ws5 = wb.create_sheet("Contact Email Sequences")
    build_sheet5_contact_sequences(ws5, data.get("contact_sequences", []))
    print(f"  ✓ Sheet 5 — Contact Email Sequences ({len(data.get('contact_sequences', []))} rows)")

    ws6 = wb.create_sheet("Territory Health")
    build_sheet6_territory_health(ws6, data.get("territory_health") or {})
    print(f"  ✓ Sheet 6 — Territory Health Dashboard")

    ws7 = wb.create_sheet("Audit Log (Manager)")
    build_sheet7_audit_log(ws7, data.get("audit_log", []))
    print(f"  ✓ Sheet 7 — Audit Log ({len(data.get('audit_log', []))} rows)")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    return output_path


if __name__ == "__main__":
    main()
